from urllib.request import urlopen
from bs4 import BeautifulSoup
import openpyxl as xl
import os
import firebase_admin
from firebase_admin import credentials
from firebase_admin import db
from selenium import webdriver
import requests
import re

path = os.path.dirname(__file__)
wb = xl.load_workbook(path+"/코로나데이터.xlsx", data_only = True)
Locate = xl.load_workbook(path+"/코로나위치좌표.xlsx", data_only = True)
driver = webdriver.Chrome(path+"/chromedriver")
places = []

status = wb["발생 현황"]
route = wb["이동 경로"]
clinic = wb["선별진료소"]
mask = wb["마스크 판매처"]
Location = Locate["좌표"]
locaterow = 0
for i in Location:
    if i[0].value != None:
        locaterow += 1
xy_url = "http://maps.googleapis.com/maps/api/geocode/xml?address="
# 카카오맵 api주소
Tbuilding = "https://apis.openapi.sk.com/tmap/jsv2?version=1&appKey=발급받은 Appkey"
building = 'https://dapi.kakao.com/v2/local/search/keyword.json?query=' # 건물명 검색
address = 'https://dapi.kakao.com/v2/local/search/address.json?query=' # 주소 검색
headers = { "Authorization": "KakaoAK 8d63652d146dc5a3a958048e957ba061"}
locatecache = {}

def addresslocation(PLACE): # 주소로 좌표획득
    global locaterow
    locaterow += 1
    if "," in PLACE:
        PLACE = PLACE.split(",")[0]
    for i in Location.rows:
        if i[0].value == PLACE:
            locaterow -= 1
            return [i[1].value, i[2].value]
    location = requests.get(address+PLACE, headers = headers).json()['documents']
    if len(location):
        location = location[0]
        Location.cell(locaterow, 1, PLACE)
        Location.cell(locaterow, 2, float(location['x']))
        Location.cell(locaterow, 3, float(location['y']))
        Locate.save(path+"/코로나위치좌표.xlsx")
        return [float(location['x']), float(location['y'])]
    locaterow -= 1
    return [0,0]

def buildinglocation(PLACE): # 건물명으로 좌표획득
    global locaterow
    locaterow += 1
    if "농산물직판장" in PLACE:
        PLACE = PLACE.split("직판장")[0]
    PLACE = "".join(PLACE.split(" "))
    for i in Location.rows:
        if i[0].value == None:
            break
        place = "".join(i[0].value.split(" "))
        if place == PLACE:
            locaterow -= 1
            return [i[1].value, i[2].value]
    location = requests.get(building+PLACE, headers = headers).json()['documents']
    if len(location):
        location = location[0]
        Location.cell(locaterow, 1, PLACE)
        Location.cell(locaterow, 2, float(location['x']))
        Location.cell(locaterow, 3, float(location['y']))
        Locate.save(path+"/코로나위치좌표.xlsx")
        return [float(location['x']), float(location['y'])]
    locaterow -= 1
    return [0,0]

def routeaddress(string):
    for i in "월화수목금토일":
        I = "(%s)"%i
        if I in string:
            string = string.split(I)[-1]
    strlist = " ".join(re.compile('[(|)|가-힣|a-z|A-Z]+').findall(string))
    strlist = re.findall("\S+", strlist)
    Strlist = []
    for i in strlist:
        if len(i)^1:
            Strlist.append(i)
    strlist = " ".join(Strlist)
    #ban = [':','(',')','월','화','수','목','금','토','일',' ']
    if "동)" in strlist:
        strlist = strlist.split("동)")[0]+"동)"
    elif "읍)" in strlist:
        strlist = strlist.split("읍)")[0]+"읍)"
    elif "면)" in strlist:
        strlist = strlist.split("면)")[0]+"면)"
    elif "방문" in strlist:
        strlist = strlist.split("방문")[0]
    elif "입원" in strlist:
        strlist = strlist.split("입원")[0]
    elif "점" in strlist:
        strlist = strlist.split("점")[0]+"점"
    if "근무" in strlist:
        strlist = "".join(strlist.split("근무"))
    if "의원" in strlist:
        if "영상의학과의원" in strlist:
            strlist = "".join(strlist.split("의원"))
        strlist = strlist
    else:
        if "내과" in strlist:
            strlist = "".join(strlist.split("내과"))
        if "정신과" in strlist:
            strlist = "".join(strlist.split("정신과"))
        if "이비인후과" in strlist:
            strlist = "".join(strlist.split("이비인후과"))
        if "외과" in strlist:
            strlist = "".join(strlist.split("외과"))
        if "응급실" in strlist:
            strlist = "".join(strlist.split("응급실"))
        if "()" in strlist:
            strlist = strlist.split("()")[0]
    if "검사" in strlist:
        strlist = "".join(strlist.split("검사"))
    if "선별진료" in strlist:
        strlist = strlist.split(" 선별진료")[0]
        if "보건소" == strlist:
            strlist = "경산시 보건소"
        elif "대구" == strlist or "대구 드라이브스루" == strlist:
            strlist = ""
    if "진료" in strlist:
        strlist = "".join(strlist.split("진료"))
    if "경산시청" in strlist:
        strlist = "경산시청"
    if "입원" in strlist:
        strlist = "".join(strlist.split("입원"))
    if '칠곡' in strlist:
        searchstr = strlist
    elif '대구' in strlist:
        searchstr = strlist
    elif '경산' in strlist:
        searchstr = strlist
    elif '김천' in strlist:
        searchstr = strlist
    elif '영천' in strlist:
        searchstr = strlist
    elif '청도' in strlist:
        searchstr = strlist
    elif '시지지' in strlist:
        searchstr = strlist
    elif '계대동산병원' in strlist:
        searchstr = "계명대학교 대구동산병원"
    elif '경북대' in strlist:
        searchstr = strlist
    elif '영대' in strlist:
        searchstr = strlist
    elif '영남대' in strlist:
        searchstr = strlist
    elif strlist == '':
        searchstr = '경산나경물류' # 검색결과가 없어서 사용
    else:
        searchstr = '경산 ' + strlist
        #searchstr = strlist
    places.append(searchstr)
    if "약국" in strlist:
        if "펜타힐즈" in strlist:
            strlist = "펜타약국"
        for i in mask.rows:
            if i[1].value in strlist:
                searchstr = i[1].value
                if searchstr in locatecache:
                    locate = locatecache[searchstr]
                    return locate
                else:
                    locate = addresslocation("경산시 "+i[4].value)
                    locatecache[searchstr] = locate
                    Location.cell(locaterow, 1, strlist)
                    Locate.save(path+"/코로나위치좌표.xlsx")
                    return locate
    if searchstr in locatecache:
        locate = locatecache[searchstr]
        return locate
    else:
        locate = buildinglocation(searchstr)
        if locate == [0, 0]:
            print("%s\n" %searchstr)
            correctname = input()
            if correctname in locatecache:
                locate = locatecache[correctname]
                return locate
            locate = buildinglocation(correctname)
            if locate != [0, 0]:
                locatecache[searchstr] = locate
                return locate
            print("%s\n" %searchstr)
            Address = input()
            if Address == "":
                locatecache[searchstr] = [0, 0]
                return [0, 0]
            if "대구" not in Address and "칠곡" not in Address and "김천" not in Address:
                Address = "경상북도 경산시 " + Address
            locate = addresslocation(Address)
            Location.cell(locaterow, 1, correctname)
            Locate.save(path+"/코로나위치좌표.xlsx")
        locatecache[searchstr] = locate
        return locate

# 확진자 이동경로 크롤링
def movingroute():
    Day = ""
    driver.get("http://www.gbgs.go.kr/programs/coronaMove/coronaMove.do")
    #html = driver.find_element_by_xpath('html/body').find_elements_by_xpath('div')[2].find_element_by_xpath('div/div/div').find_elements_by_xpath('table')
    html = driver.page_source
    bsObject = BeautifulSoup(html, "html.parser").find_all("table")
    Patients = {}
    row = 1
    for i in bsObject:
        I = i.tbody
        Info=I.tr
        info = []
        I = I.contents
        for j in Info:
            t=str(type(j))
            if "bs4.element.NavigableString" not in t and "bs4.element.Comment" not in t:
                info.append(j)
        route.cell(row, 1, int(info[0].contents[0])) # 경산시 내부 번호
        if info[1].contents[0] == 0:
            route.cell(row, 2, "미정") # 확진 번호
        else:
            route.cell(row, 2, int(info[1].contents[0])) # 확진 번호
        route.cell(row, 3, info[2].contents[0]) # 인적사항 : 성별(거주하는 동, 나이)
        day = info[3].contents[0].split(".")
        route.cell(row, 4, day[0]+"월 "+day[1]+"일")
        if len(info[4]) == 0:
            route.cell(row, 5, "배정요청")
        elif len(info[4].contents) > 1:
            center = []
            for j in info[4].contents:
                t=str(type(j))
                if "bs4.element.NavigableString" in t or "bs4.element.Comment" in t:
                    center.append(j)
            route.cell(row, 5, "\n".join(center))
        else:
            route.cell(row, 5, info[4].contents[0])
        if len(info[5]) == 0:
            route.cell(row, 6, '확인중(확인중)')
        else:
            route.cell(row, 6, info[5].contents[0])
        where = I[3].find_all("li")
        extra = []
        for j in where:
            if j.contents[0] != "\n":
                j = j.contents[0]
            else:
                continue
            if "경로 확인중" in j:
                route.cell(row, 7, "경로 확인중")
            elif "주요동선(직장, 병의원, 약국) 없음" in j:
                route.cell(row, 7, "주요동선(직장, 병의원, 약국) 없음")
            else:
                if "※" in j:
                    i = 0
                    while j[i] == " ":
                        i += 1
                    if j[i] == "※":
                        extra.append(j)
                    else:
                        j = j.split("※")
                        extra.append("※"+j[1])
                        j = j[0]
                        jlist = [j]
                        if "\n" in j:
                            jlist = j.split("\n")
                        for k in jlist:
                            if k == "":
                                continue
                            row += 1
                            if "퇴원" not in k and "사망" not in k and "격리해제" not in k and "자택" not in k and "신천지" not in k and "직장" not in k and "양성" not in k and "자가격리" not in k and "없음" not in k and "두통" not in k and "발열" not in k and "근육통" not in k and "나경물류" not in k and "한국아이피엠" not in k:
                                locate = routeaddress(k)
                                if locate != [0, 0]:
                                    route.cell(row, 8, float(locate[0]))
                                    route.cell(row, 9, float(locate[1]))
                            route.cell(row, 7, k)
                else:
                    if j == "":
                        continue
                    row += 1
                    day = []
                    J = j
                    for i in "월화수목금토일":
                        I = "(%s)"%i
                        if I in J:
                            day.append(J.split(I)[0]+I)
                            J = J.split(I)[-1]
                    if len(day):
                        if ":" in J:
                            N = J.index(":")
                            day.append(J[:N+3])
                        Day = "".join(day)
                    else:
                        j = Day + j
                    if "퇴원" not in j and "사망" not in j and "격리해제" not in j and "자택" not in j and "신천지" not in j and "직장" not in j and "양성" not in j and "자가격리" not in j and "없음" not in j and "두통" not in j and "발열" not in j and "근육통" not in j and "나경물류" not in j and "한국아이피엠" not in j:
                        locate = routeaddress(j)
                        if locate != [0, 0]:
                            route.cell(row, 8, float(locate[0]))
                            route.cell(row, 9, float(locate[1]))
                    route.cell(row, 7, j)
                for k in extra:
                    row += 1
                    route.cell(row, 10, k)
        row += 1
    wb.save(path+"/코로나데이터.xlsx")

def occurrence():
    html = urlopen("http://www.gbgs.go.kr/programs/corona/corona.do")
    gs = BeautifulSoup(html, "html.parser").body.contents[5].div.div.find_all('div')[1] # 경산데이터
    updatetime = gs.contents[3].contents[0].split("기준")[0].split(".")
    updatetime = "20%s년 %s월 %s일 %s 기준" %(updatetime[0], updatetime[1], updatetime[2], updatetime[3][1:])
    status.cell(1, 1, updatetime) # 업데이트 시간
    info = gs.contents[5].contents
    patients = info[1].ul.find_all("li") # 확진자 데이터
    diagnosis = info[5].ul.find_all("li") # 의심환자 데이터
    status.cell(2, 1, "전체")
    status.cell(2, 2, int("".join(patients[0].span.contents[0].split(",")))) # 확진자 수
    status.cell(2, 3, int("".join(diagnosis[1].contents[1].contents[0].split(",")))) # 검사중인 환자 수
    status.cell(2, 4, int("".join(patients[3].span.contents[0].split(",")))) # 퇴원자 수
    status.cell(2, 5, int("".join(diagnosis[2].span.contents[0].split(",")))) # 음성인 환자 수
    area = []
    for i in route.rows:
        if i[2].value != None:
            I = i[2].value.split("(")[1].split(",")[0]
            if I.isdigit():
                I = i[2].value.split("(")[1].split(",")[1].split(")")[0]
            if I not in area:
                area.append(I)
    area = sorted(area)
    Area = {}
    row = 2
    for i in area:
        row += 1
        Area[i] = 0
        status.cell(row, 1, i) # 동/읍/면
        status.cell(row, 2, "0") # 동/읍/면 별 확진자 수 초기화
    for i in route.rows:
        if i[2].value != None:
            i = i[2].value.split("(")[1].split(",")[0]
            r = 2
            for j in area:
                r += 1
                if j == i:
                    status.cell(r, 2, str(int(status.cell(r, 2).value) + 1)) # 동/읍/면 별 확진자 수 초기화
                    break
    for i in status.rows:
        if i[1].value != None:
            i[1].value = int(i[1].value)
    wb.save(path+"/코로나데이터.xlsx")

def maskinfo():
    html = urlopen("http://www.gbgs.go.kr/design/health/COVID19/COVID19_05_02.html") # 경산시 마스크 공적판매처
    OfficialMask = BeautifulSoup(html, "html.parser").body.contents[5].div.div.div.div.contents
    OfficialMask =[OfficialMask[9].table.tbody, OfficialMask[15].table.tbody]
    Dong = []
    dong = OfficialMask[0].contents
    for i in dong:
        t=str(type(i))
        if "bs4.element.NavigableString" not in t and "bs4.element.Comment" not in t:
            Dong.append(i)
    EupMyeon = []
    eupmyeon = OfficialMask[1].contents
    for i in eupmyeon:
        t=str(type(i))
        if "bs4.element.NavigableString" not in t and "bs4.element.Comment" not in t:
            EupMyeon.append(i)
    row = 1
    for i in Dong: # 동지역 판매처
        I = []
        for j in i:
            t=str(type(j))
            if "bs4.element.NavigableString" not in t and "bs4.element.Comment" not in t:
                I.append(j)
        if len(I) == 3:
            mask.cell(row, 1, I[0].contents[0].split("\\")[0]) # 분류
            mask.cell(row, 2, I[1].contents[0]) # 판매처 이름
            mask.cell(row, 3, I[2].contents[0]) # 전화번호
            mask.cell(row, 4, "공적판매처") # 공사구분
            mask.cell(row, 5, "동") # 동 읍/면 구분
        else:
            mask.cell(row, 2, I[0].contents[0]) # 판매처 이름
            mask.cell(row, 3, I[1].contents[0]) # 전화번호
            mask.cell(row, 4, "공적판매처") # 공사구분
            mask.cell(row, 5, "동") # 동 읍/면 구분
        row += 1
    for i in EupMyeon: # 읍면지역 판매처
        I = []
        for j in i:
            t=str(type(j))
            if "bs4.element.NavigableString" not in t and "bs4.element.Comment" not in t:
                I.append(j)
        if len(I) == 3:
            mask.cell(row, 1, I[0].contents[0].split("\\")[0]) # 분류
            place = I[1].contents[0] # 판매처 이름
            mask.cell(row, 3, I[2].contents[0]) # 전화번호
        else:
            place = I[0].contents[0] # 판매처 이름
            mask.cell(row, 3, I[1].contents[0]) # 전화번호
        if "우체국" in place:
            mask.cell(row, 2, "경산"+place)
        elif "농협" in place:
            if "하나로마트" in place:
                Place = place.split("하나로마트")
                if Place[1] == "":
                    if " " in Place[0]:
                        Place[0] = Place[0].split(" ")
                        if Place[0][0] != "":
                            Place[0] = Place[0][0]
                        else:
                            Place[0] = Place[0][1]
                    if "농협" in Place[0]:
                        Place = Place[0].split("농협")
                    if "지점" in Place[1]:
                        Place[1] = Place[1].split("지점")[0]+"점"
                    mask.cell(row, 2, Place[0]+"농협 하나로마트 "+Place[1])
                elif Place[0] == "":
                    if " " in Place[1]:
                        Place[1] = Place[1].split(" ")
                        if Place[1][0] != "":
                            Place[1] = Place[1][0]
                        else:
                            Place[1] = Place[1][1]
                    if "농협" in Place[1]:
                        Place = Place[1].split("농협")
                    if "지점" in Place[1]:
                        Place[1] = Place[1].split("지점")[0]+"점"
                    mask.cell(row, 2, Place[0]+"농협 하나로마트 "+Place[1])
                else:
                    if " " in Place[0]:
                        Place[0] = Place[0].split(" ")
                        if Place[0][0] != "":
                            Place[0] = Place[0][0]
                        else:
                            Place[0] = Place[0][1]
                    if "농협" in Place[0]:
                        Place1 = Place[0].split("농협")
                    if "지점" in Place[1]:
                        Place[1] = Place[1].split("지점")[0]+"점"
                    mask.cell(row, 2, Place1[0]+"농협 하나로마트 "+Place[1])
            else:
                mask.cell(row, 2, place)
        mask.cell(row, 4, "공적판매처") # 공사구분
        mask.cell(row, 5, "읍/면") # 동 읍/면 구분
        row += 1
    html = urlopen("http://www.gbgs.go.kr/design/health/COVID19/COVID19_05_03.html") # 경산시 마스크 판매 약국
    pharmacy = BeautifulSoup(html, "html.parser").body.contents[5].div.div.table.tbody.contents
    for i in pharmacy[1::2]:
        i = i.contents
        mask.cell(row, 1, i[3].contents[0]) # 읍면동
        mask.cell(row, 2, i[5].contents[0]) # 판매처이름
        mask.cell(row, 3, i[9].contents[0]) # 전화번호
        mask.cell(row, 4, "약국") # 공사구분
        mask.cell(row, 5, i[7].contents[0]) # 주소
        row += 1
    driver.get("http://www.gbgs.go.kr/design/health/COVID19/COVID19_05_05.html") # 경산시 판매 현황
    html = driver.page_source
    stock = BeautifulSoup(html, "html.parser").body.contents[5].div.div.div.div.table.tbody.contents
    for i in stock[1::]:
        row = 1
        i = i.contents
        name = "".join(i[0].contents[0].split("본점")) # 판매처 이름
        name = "".join(name.split(" "))
        if mask.cell(row, 2).value != None:
            Name = "".join(mask.cell(row, 2).value.split(" "))
        else:
            Name = None
        while Name not in [None,name]: # 판매처 인덱스 찾기
            row += 1
            if mask.cell(row, 2).value != None:
                Name = "".join(mask.cell(row, 2).value.split(" "))
            else:
                row -= 1
                Name = None
        if len(i[2].contents) == 0:
            Stock = "없음"
        else:
            Stock = i[2].contents[0]
        if "(" in Stock:
            Stock = Stock.split(" (")
            Stock[1] = "(" + Stock[1]
        else:
            Stock = [Stock, ""]
        mask.cell(row, 6, Stock[0]) # 재고 수준
        mask.cell(row, 7, Stock[1]) # 재고량
        if len(i[4].contents):
            renewaltime = i[4].contents[0]
        mask.cell(row, 8, renewaltime) # 재고량 갱신 시간
    row = 0
    for i in mask.rows:
        row += 1
        if i[1].value == None:
            break
        elif i[5].value == None:
            i[5].value = "집계중"
            i[7].value = renewaltime
        if i[3].value == "공적판매처":
            place = i[1].value
            #if "하나로마트" in place:
                #if place[-1] != "점":
                #    place += "본점"
            locate = buildinglocation(place)
            mask.cell(row, 9, locate[0])
            mask.cell(row, 10, locate[1])

        elif i[3].value == "약국":
            place = "경산시 "+i[4].value
            locate = addresslocation(place)
            mask.cell(row, 9, locate[0])
            mask.cell(row, 10, locate[1])

    wb.save(path+"/코로나데이터.xlsx")

def clinicinfo(): # 선별진료소 데이터 크롤링
    html = urlopen("http://www.gbgs.go.kr/design/health/COVID19/COVID19_04.html") # 선별 진료소 현황
    Clinic = BeautifulSoup(html, "html.parser").body.contents[5].div.div.div.table.tbody.contents
    row = 1
    for i in Clinic[1::2]:
        i = i.contents
        clinic.cell(row, 1, i[1].contents[0]) # 기관
        clinic.cell(row, 2, i[3].contents[0]) # 주소
        call = []
        for j in i[5].contents:
            t=str(type(j))
            if "bs4.element.NavigableString" in t or "bs4.element.Comment" in t:
                j = j.split("\r")
                j = j[-1].split("\n")
                j = j[-1].split("\t")[-1]
                call.append(j)
        clinic.cell(row, 3, " \r\n".join(call)) # 전화번호
        time = []
        for j in i[7].contents:
            t=str(type(j))
            if "bs4.element.NavigableString" in t or "bs4.element.Comment" in t:
                j = j.split("\r")
                j = j[-1].split("\n")
                j = j[-1].split("\t")[-1]
                time.append(j)
        clinic.cell(row, 4, " \r\n".join(time)) # 진료시간
        if len(i) > 9:
            extra = []
            I = i[9].contents
            if len(I) == 1 and "bs4.element.Tag" in str(type(I[0])):
                I = I[0].contents
            for j in I:
                t=str(type(j))
                if "bs4.element.NavigableString" in t or "bs4.element.Comment" in t:
                    j = j.split("\r")
                    j = j[-1].split("\n")
                    j = j[-1].split("\t")[-1]
                    extra.append(j)
            clinic.cell(row, 5, " \r\n".join(extra))
        else:
            extra = clinic.cell(row-1, 5).value
            clinic.cell(row, 5, extra)
        locate = addresslocation("경산시 " + clinic.cell(row, 2).value)
        clinic.cell(row, 6, locate[0])
        clinic.cell(row, 7, locate[1])
        row += 1
    wb.save(path+"/코로나데이터.xlsx")

def crawler():
    movingroute()
    occurrence()
    clinicinfo()
    maskinfo()

def firebase(): # 파이어베이스 업로드
    cred = credentials.Certificate(path+'/covid19-daegu-gyeongsan-firebase-adminsdk-wfe8g-097e7accda.json')
    firebase_admin.initialize_app(cred,{
        'databaseURL' : 'https://covid19-daegu-gyeongsan.firebaseio.com/'
    })
    ref = db.reference('코로나/경산/')
    Occur = {'갱신시간':status.cell(1,1).value, '검사중':status.cell(2,3).value, '완치':status.cell(2,4).value, '음성':status.cell(2,5).value}
    Route = {}
    Clinic = {}
    Mask = {}
    r = 0
    for i in status.rows:
        if r:
            Occur[i[0].value] = i[1].value
        r += 1
    
    for i in route.rows:
        if i[0].value != None:
            Num = i[0].value
            info = {"확진번호":i[1].value, "인적사항":i[2].value, "확진일자":i[3].value, "입원기관":i[4].value, "접촉자수(격리조치중)":i[5].value, "비고" : [], "이동경로":{}}
            routenumber = 0
        elif i[6].value != None:
            routenumber += 1
            info["이동경로"][routenumber] = {"경로":i[6].value}
            if i[7].value != None:
                info["이동경로"][routenumber]["좌표"] = [i[7].value, i[8].value]
        elif i[9].value != None:
            info["비고"].append(i[9].value)
        Route[Num] = info

    for i in clinic.rows:
        if i[0].value != None:
            Clinic[i[0].value] = {"주소":i[1].value, "전화번호":i[2].value, "진료시간":i[3].value, "비고":i[4].value, "좌표":[i[5].value, i[6].value]}
    
    for i in mask.rows:
        if i[0].value != None and i[0].value != "축협":
            area = i[0].value
            mask[area] = {"공적판매처":{}, "약국":{}}
        if i[3].value == "공적판매처":
            mask[area]["공적판매처"][i[1].value] = {"전화번호":i[2].value, "재고수준":i[5].value, "갱신일시": i[7].value, "좌표": [i[8].value, i[9].value]}
            if i[6].value != None:
                mask[area]["공적판매처"][i[1].value]["재고량"] = i[6].value
        elif i[3].value == "약국":
            mask[area]["약국"][i[1].value] = {"전화번호":i[2].value, "주소": i[4].value, "재고수준":i[5].value, "갱신일시": i[7].value, "좌표": [i[8].value, i[9].value]}
            if i[6].value != None:
                mask[area]["공적판매처"][i[1].value]["재고량"] = i[6].value

crawler()